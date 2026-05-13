#!/usr/bin/env python3
"""
Ingest grave photos from the inbox, update the AD_Master xlsx file_name column,
regenerate the amalgamated register CSV, and write it into the website's
assets/list folder where the Eleventy build picks it up.

Usage:
    python scripts/ingest_and_rebuild.py              # Full run: ingest + rebuild
    python scripts/ingest_and_rebuild.py --dry-run    # Preview only; no files touched
    python scripts/ingest_and_rebuild.py --skip-inbox # Rebuild CSV from current xlsx only

Every run also reconciles the website's existing graves folder against the xlsx:
photos already on disk whose grave row has an empty file_name get auto-linked,
and any file_name that points to a missing photo is reported as a broken link.

Source files (both checked into the repo under assets/list/):
  - AD_Master_Data Base Templecurraheen Graveyard Feb23.xlsx  (stones data)
  - Old_Templecurraheen_Register_of_deaths.txt                (civil register)
"""

import argparse
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
ASSETS_LIST_DIR = REPO_ROOT / "assets" / "list"

XLSX_PATH = ASSETS_LIST_DIR / "AD_Master_Data Base Templecurraheen Graveyard Feb23.xlsx"
REGISTER_TXT_PATH = ASSETS_LIST_DIR / "Old_Templecurraheen_Register_of_deaths.txt"
CSV_OUTPUT_PATH = ASSETS_LIST_DIR / "templecurraheen_stones_and_registers.csv"

INBOX_DIR = Path(r"D:\Templecurraheen\photos-inbox")
WEBSITE_GRAVES_DIR = REPO_ROOT / "assets" / "imgs" / "graves"

BACKUPS_DIR = SCRIPT_DIR / "backups"
REPORTS_DIR = SCRIPT_DIR / "ingest_reports"

PHOTO_PATTERN = re.compile(r'^([a-z])(-?\d+)([a-z]*)\.(jpg|jpeg|png)$', re.IGNORECASE)
IMAGE_EXTS = {'.jpg', '.jpeg', '.png'}


def ts():
    return datetime.now().strftime('%Y%m%d_%H%M%S')


def scan_inbox():
    if not INBOX_DIR.exists():
        return []
    photos = []
    for p in sorted(INBOX_DIR.iterdir()):
        if not p.is_file() or p.suffix.lower() not in IMAGE_EXTS:
            continue
        m = PHOTO_PATTERN.match(p.name)
        if m:
            row, grave, plot, _ = m.groups()
            photos.append((p, {
                'row': row.upper(),
                'grave': grave,
                'plot': plot.lower(),
            }))
        else:
            photos.append((p, None))
    return photos


def find_matching_rows(ws, parsed, headers):
    gl_col = headers['Grave Location']
    pc_col = headers.get('plot_code')
    matches = []
    for row_idx in range(2, ws.max_row + 1):
        gl_val = ws.cell(row=row_idx, column=gl_col).value
        if gl_val is None:
            continue
        gl_str = str(gl_val).strip()
        if not gl_str:
            continue
        row_letter = gl_str[0].upper()
        digits = re.sub(r'[^0-9\-]', '', gl_str)
        pc_val = ws.cell(row=row_idx, column=pc_col).value if pc_col else None
        pc_str = str(pc_val).strip().lower() if pc_val is not None and str(pc_val).strip() else ''
        if (row_letter == parsed['row']
                and digits == parsed['grave']
                and pc_str == parsed['plot']):
            matches.append(row_idx)
    return matches


def update_xlsx_file_names(photos, dry_run):
    """
    Mutates the xlsx in place (unless dry_run).
    Returns (update_log, overwrite_log, no_match_log, matched_filenames).
    """
    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val).strip()] = col_idx

    for required in ('Grave Location', 'file_name'):
        if required not in headers:
            raise RuntimeError(
                f"Required column '{required}' not found in xlsx. "
                f"Headers seen: {sorted(headers.keys())}"
            )

    fn_col = headers['file_name']

    update_log = []
    overwrite_log = []
    no_match_log = []
    matched_filenames = set()

    for photo_path, parsed in photos:
        if parsed is None:
            no_match_log.append((photo_path.name, 'filename does not match {row}{grave}{plot?}.ext pattern'))
            continue

        matches = find_matching_rows(ws, parsed, headers)
        if not matches:
            no_match_log.append((
                photo_path.name,
                f"no xlsx row for row={parsed['row']}, grave={parsed['grave']}, plot={parsed['plot'] or '<none>'}"
            ))
            continue

        for row_idx in matches:
            existing = ws.cell(row=row_idx, column=fn_col).value
            existing_str = str(existing).strip() if existing is not None else ''
            if existing_str and existing_str != photo_path.name:
                overwrite_log.append((photo_path.name, row_idx, existing_str))
            if not dry_run:
                ws.cell(row=row_idx, column=fn_col).value = photo_path.name

        update_log.append((photo_path.name, len(matches)))
        matched_filenames.add(photo_path.name)

    if not dry_run and update_log:
        wb.save(XLSX_PATH)

    return update_log, overwrite_log, no_match_log, matched_filenames


def rebuild_csv(output_path):
    """Replicates Amalgamate_Templecurraheen_Graveyard_Data.ipynb exactly."""
    tc_df = pd.read_excel(XLSX_PATH)

    tc_df['grave_code'] = tc_df['Grave Location'].str.replace(r'[^0-9\-]', '', regex=True)
    tc_df['row_code'] = tc_df['Grave Location'].str[0]
    tc_df['last_name'] = tc_df['Surname']
    tc_df['first_name'] = tc_df['First Name']
    tc_df['death_date'] = tc_df['Date of Death']
    tc_df['age'] = tc_df['Age']
    tc_df['names_on_headstone'] = tc_df['Names on headstone']
    tc_df['stone_inscription'] = tc_df['Inscription on Headstone     ']
    tc_df['stone_design'] = tc_df['Stone design']
    tc_df['headstone_shape'] = tc_df['   Shape of Headstone']

    tc_cols = ['grave_code', 'row_code', 'plot_code', 'last_name', 'first_name',
               'death_date', 'age', 'names_on_headstone', 'file_name',
               'stone_inscription', 'stone_design', 'headstone_shape']
    tc_df_cleaned = tc_df[tc_cols].copy()
    tc_df_cleaned = tc_df_cleaned.dropna(subset=['grave_code'])

    tc_df_cleaned['grave_code_first_char'] = tc_df_cleaned['grave_code'].str[0]
    tc_df_cleaned['is_uppercase_alpha'] = tc_df_cleaned['grave_code_first_char'].str.match(r'^[A-Z]$')

    df_ch = pd.read_csv(
        REGISTER_TXT_PATH,
        sep=',',
        encoding='utf-8',
        quotechar='"',
        engine='python',
        skipinitialspace=True,
        on_bad_lines=lambda x: x[:14],
    )
    df_ch['grave_code'] = "50" + df_ch['Item'].astype(str)
    df_ch['last_name'] = df_ch['Last_Name']
    df_ch['first_name'] = df_ch['First_Name']
    df_ch['age'] = df_ch['Age']
    df_ch['death_date'] = df_ch['Death_Date']
    df_ch['row_code'] = "X"
    df_ch['plot_code'] = " "
    df_ch['register_page'] = df_ch['Page']

    df_ch_cleaned = df_ch[['grave_code', 'row_code', 'plot_code', 'last_name',
                           'first_name', 'age', 'death_date', 'Burial_Date',
                           'Months', 'Status', 'Occupation', 'Address',
                           'Witness', 'Comments', 'register_page']].copy()

    combined = pd.concat([tc_df_cleaned, df_ch_cleaned], ignore_index=True)

    try:
        combined['grave_code'] = combined['grave_code'].astype(int)
    except (ValueError, TypeError) as e:
        print(f"WARNING: grave_code int-cast failed ({e}); falling back to string sort.")

    combined_sorted = combined.sort_values('grave_code', kind='stable')
    combined_sorted.to_csv(output_path, index=False)
    return len(combined_sorted)


def scan_graves_dir():
    """Scan the website's graves folder. Returns list of (path, parsed-or-None)."""
    if not WEBSITE_GRAVES_DIR.exists():
        return []
    photos = []
    for p in sorted(WEBSITE_GRAVES_DIR.iterdir()):
        if not p.is_file() or p.suffix.lower() not in IMAGE_EXTS:
            continue
        m = PHOTO_PATTERN.match(p.name)
        if m:
            row, grave, plot, _ = m.groups()
            photos.append((p, {
                'row': row.upper(),
                'grave': grave,
                'plot': plot.lower(),
            }))
        else:
            photos.append((p, None))
    return photos


def reconcile_existing_photos(dry_run):
    """
    Walk every photo already in WEBSITE_GRAVES_DIR and link it to its xlsx row
    when the row's file_name cell is empty. Never overwrite a non-empty cell -
    those are flagged for manual review. Also flag any non-empty file_name that
    points to a photo no longer on disk.

    Returns (linked, conflicts, no_match, unparsed, broken_links).
    """
    photos = scan_graves_dir()
    on_disk = {p.name for p, _ in photos}

    wb = load_workbook(XLSX_PATH)
    ws = wb.active

    headers = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            headers[str(val).strip()] = col_idx
    fn_col = headers['file_name']

    linked = []          # (photo_name, row_idx)
    conflicts = []       # (photo_name, row_idx, existing_value)
    no_match = []        # (photo_name, reason)
    unparsed = []        # (photo_name,)
    broken_links = []    # (row_idx, missing_file_name, surname)

    sur_col = headers.get('Surname')

    for photo_path, parsed in photos:
        if parsed is None:
            unparsed.append(photo_path.name)
            continue
        matches = find_matching_rows(ws, parsed, headers)
        if not matches:
            no_match.append((photo_path.name,
                             f"no xlsx row for row={parsed['row']}, grave={parsed['grave']}, plot={parsed['plot'] or '<none>'}"))
            continue
        for row_idx in matches:
            existing = ws.cell(row=row_idx, column=fn_col).value
            existing_str = str(existing).strip() if existing is not None else ''
            if not existing_str:
                if not dry_run:
                    ws.cell(row=row_idx, column=fn_col).value = photo_path.name
                linked.append((photo_path.name, row_idx))
            elif existing_str == photo_path.name:
                continue  # already linked, nothing to do
            else:
                conflicts.append((photo_path.name, row_idx, existing_str))

    # Broken-link sweep: any non-empty file_name pointing at a missing file
    for row_idx in range(2, ws.max_row + 1):
        existing = ws.cell(row=row_idx, column=fn_col).value
        existing_str = str(existing).strip() if existing is not None else ''
        if not existing_str:
            continue
        if existing_str not in on_disk:
            surname = ''
            if sur_col:
                sv = ws.cell(row=row_idx, column=sur_col).value
                surname = str(sv).strip() if sv is not None else ''
            broken_links.append((row_idx, existing_str, surname))

    if not dry_run and linked:
        wb.save(XLSX_PATH)

    return linked, conflicts, no_match, unparsed, broken_links


def move_matched_photos(photos, matched_filenames, dry_run):
    moved = []
    skipped = []
    for photo_path, parsed in photos:
        if photo_path.name not in matched_filenames:
            skipped.append(photo_path.name)
            continue
        dest = WEBSITE_GRAVES_DIR / photo_path.name
        if dry_run:
            moved.append((photo_path.name, 'would move'))
        else:
            shutil.move(str(photo_path), str(dest))
            moved.append((photo_path.name, 'moved'))
    return moved, skipped


def main():
    parser = argparse.ArgumentParser(description='Ingest grave photos and rebuild the register CSV.')
    parser.add_argument('--dry-run', action='store_true', help='Preview only; do not modify any files.')
    parser.add_argument('--skip-inbox', action='store_true', help='Skip photo ingestion; rebuild CSV only.')
    parser.add_argument('--skip-reconcile', action='store_true',
                        help='Skip the website graves-folder reconciliation step.')
    args = parser.parse_args()

    stamp = ts()
    log = [f"# Ingest Report - {stamp}", "",
           f"- Dry run: **{args.dry_run}**",
           f"- Skip inbox: **{args.skip_inbox}**",
           f"- Skip reconcile: **{args.skip_reconcile}**"]

    will_mutate_xlsx = (not args.dry_run) and (not args.skip_inbox or not args.skip_reconcile)
    if will_mutate_xlsx:
        BACKUPS_DIR.mkdir(exist_ok=True)
        backup_path = BACKUPS_DIR / f"AD_Master_{stamp}.xlsx"
        shutil.copy2(XLSX_PATH, backup_path)
        log.append(f"- Backup: `{backup_path.name}`")
    else:
        log.append("- Backup: _skipped_")

    matched_filenames = set()
    photos = []

    if args.skip_inbox:
        log.append("\n## Photo ingestion: SKIPPED")
    else:
        photos = scan_inbox()
        log.append(f"\n## Photo ingestion\n\nInbox: `{INBOX_DIR}` - {len(photos)} image(s) found.")

        if photos:
            update_log, overwrite_log, no_match_log, matched_filenames = update_xlsx_file_names(photos, args.dry_run)

            if update_log:
                log.append(f"\n### Mapped successfully ({len(update_log)})")
                for name, n_rows in update_log:
                    log.append(f"- `{name}` -> {n_rows} row(s) updated")

            if overwrite_log:
                log.append(f"\n### Overwrote previous mapping ({len(overwrite_log)})")
                for new, row_idx, old in overwrite_log:
                    log.append(f"- xlsx row {row_idx}: `{old}` -> `{new}`")

            if no_match_log:
                log.append(f"\n### Left in inbox ({len(no_match_log)})")
                for name, reason in no_match_log:
                    log.append(f"- `{name}`: {reason}")

    if args.skip_reconcile:
        log.append("\n## Website graves reconcile: SKIPPED")
    else:
        linked, conflicts, no_match_g, unparsed, broken = reconcile_existing_photos(args.dry_run)
        log.append(f"\n## Website graves reconcile\n\nFolder: `{WEBSITE_GRAVES_DIR}`")

        if linked:
            verb = "Would link" if args.dry_run else "Linked"
            log.append(f"\n### {verb} ({len(linked)})")
            for name, row_idx in linked:
                log.append(f"- `{name}` -> xlsx row {row_idx}")

        if conflicts:
            log.append(f"\n### Conflicts - manual review ({len(conflicts)})")
            log.append("_Photo on disk matches the row by code, but the row already lists a different file. Not overwritten._")
            for name, row_idx, old in conflicts:
                log.append(f"- xlsx row {row_idx}: keeps `{old}`; on disk: `{name}`")

        if no_match_g:
            log.append(f"\n### Photos with no matching xlsx row ({len(no_match_g)})")
            for name, reason in no_match_g:
                log.append(f"- `{name}`: {reason}")

        if unparsed:
            log.append(f"\n### Photos with unparseable filenames ({len(unparsed)})")
            for name in unparsed:
                log.append(f"- `{name}`")

        if broken:
            log.append(f"\n### Broken file_name links ({len(broken)})")
            log.append("_xlsx row references a photo that is not on disk. Not auto-cleared._")
            for row_idx, missing, surname in broken:
                tag = f" ({surname})" if surname else ""
                log.append(f"- xlsx row {row_idx}{tag}: `{missing}` is missing")

    if args.dry_run:
        dry_csv = ASSETS_LIST_DIR / f"templecurraheen_stones_and_registers.DRYRUN_{stamp}.csv"
        row_count = rebuild_csv(dry_csv)
        log.append(f"\n## CSV rebuild (dry run)\n\n- Output: `{dry_csv.name}`\n- Rows: **{row_count}**")
    else:
        row_count = rebuild_csv(CSV_OUTPUT_PATH)
        log.append(f"\n## CSV rebuild\n\n- Output: `{CSV_OUTPUT_PATH}`\n- Rows: **{row_count}**")

    if not args.skip_inbox and photos and matched_filenames:
        moved, skipped = move_matched_photos(photos, matched_filenames, args.dry_run)
        log.append(f"\n## Photo move\n\n- Moved: {len(moved)}")
        for name, action in moved:
            log.append(f"  - `{name}` - {action}")
        if skipped:
            log.append(f"- Left in inbox: {len(skipped)}")

    REPORTS_DIR.mkdir(exist_ok=True)
    report_path = REPORTS_DIR / f"ingest_{stamp}.md"
    report_text = '\n'.join(log)
    report_path.write_text(report_text, encoding='utf-8')

    safe_text = report_text.encode('ascii', errors='replace').decode('ascii')
    print(safe_text)
    print(f"\nReport: {report_path}")


if __name__ == '__main__':
    main()
