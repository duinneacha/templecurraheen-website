"""Find graves that genuinely need a photo, accounting for photos on disk
that match the grave's row/code/plot but aren't linked in the register.

Output: assets/list/graves_missing_photos.xlsx with two sheets:
  - "Needs Photo"   — graves with no matching image on disk by any rule
  - "Link Fix Only" — graves where a correctly-named photo IS on disk but
                      the file_name column is empty or points elsewhere
And a third sheet "Orphan Photos" listing files on disk that don't match
any grave in the register.

Row X (register-only entries from the civil register) are excluded since
they have no physical stone to photograph.
"""
import csv
import os
import re
from collections import OrderedDict, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = os.path.join(os.path.dirname(__file__), "..")
CSV_PATH = os.path.join(ROOT, "assets", "list", "templecurraheen_stones_and_registers.csv")
GRAVES_DIR = os.path.join(ROOT, "assets", "imgs", "graves")
OUT_PATH = os.path.join(ROOT, "assets", "list", "graves_missing_photos.xlsx")

# Same shape as the ingest script — keep in sync if that ever changes.
PHOTO_PATTERN = re.compile(r'^([a-z])(-?\d+)([a-z]*)\.(jpg|jpeg|png)$', re.IGNORECASE)


def index_disk_photos():
    """Return (filenames_lower, by_key) where by_key maps (ROW, grave_no, plot_lower) -> [filenames]."""
    filenames = {f for f in os.listdir(GRAVES_DIR) if os.path.isfile(os.path.join(GRAVES_DIR, f))}
    available_lower = {f.lower() for f in filenames}
    by_key = defaultdict(list)
    for f in filenames:
        m = PHOTO_PATTERN.match(f)
        if not m:
            continue
        row_letter, grave_digits, plot_letter, _ext = m.groups()
        by_key[(row_letter.upper(), grave_digits, plot_letter.lower())].append(f)
    return available_lower, by_key


def main():
    available_lower, photos_by_key = index_disk_photos()

    # Group register rows by (row_code, grave_code, plot_code) — one entry per grave
    graves = OrderedDict()
    with open(CSV_PATH, "r", encoding="utf-8", newline="") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            code = (row.get("grave_code") or "").strip()
            row_code = (row.get("row_code") or "").strip()
            plot_code = (row.get("plot_code") or "").strip()
            file_name = (row.get("file_name") or "").strip()
            last_name = (row.get("last_name") or "").strip()
            first_name = (row.get("first_name") or "").strip()
            full_name = " ".join(p for p in (first_name, last_name) if p).strip()

            key = (row_code, code, plot_code)
            entry = graves.setdefault(key, {
                "row_code": row_code,
                "grave_code": code,
                "plot_code": plot_code,
                "file_names": set(),
                "names": [],
            })
            if file_name:
                entry["file_names"].add(file_name)
            if full_name:
                entry["names"].append(full_name)

    needs_photo = []
    link_fix = []
    matched_disk_files = set()  # any filename on disk that matched some grave (by either rule)

    for key, entry in graves.items():
        row_code = entry["row_code"]
        grave_code = entry["grave_code"]
        plot_code = entry["plot_code"]

        # Skip register-only X rows — no physical stone to photograph
        if row_code.upper() == "X":
            # Still mark any file_name they list as "matched" so it doesn't show as orphan
            for f in entry["file_names"]:
                if f.lower() in available_lower:
                    matched_disk_files.add(f.lower())
            continue

        # Rule 1: any listed file_name that exists on disk
        listed_present = [f for f in entry["file_names"] if f.lower() in available_lower]
        for f in listed_present:
            matched_disk_files.add(f.lower())

        # Rule 2: any disk photo whose parsed (row, grave, plot) matches this grave
        disk_match_key = (row_code.upper(), grave_code, plot_code.lower())
        disk_matches = photos_by_key.get(disk_match_key, [])
        for f in disk_matches:
            matched_disk_files.add(f.lower())

        if listed_present:
            continue  # already shown on the website — not missing

        record = {
            "row_code": row_code,
            "grave_code": grave_code,
            "plot_code": plot_code,
            "names": "; ".join(entry["names"]),
            "listed_file_name": "; ".join(sorted(entry["file_names"])),
            "disk_match": "; ".join(sorted(disk_matches)),
        }
        if disk_matches:
            link_fix.append(record)
        else:
            needs_photo.append(record)

    # Orphan photos: on disk, parseable, but no matching grave in register
    orphans = []
    register_keys = {(rc.upper(), gc, pc.lower()) for (rc, gc, pc) in graves.keys()}
    for key, files in photos_by_key.items():
        if key in register_keys:
            continue
        row_letter, grave_digits, plot_letter = key
        for f in files:
            orphans.append({
                "filename": f,
                "row_letter": row_letter,
                "grave": grave_digits,
                "plot": plot_letter,
            })

    # Unparseable files on disk (don't match the row/grave/plot pattern at all)
    unparseable = []
    for f in sorted(os.listdir(GRAVES_DIR)):
        full = os.path.join(GRAVES_DIR, f)
        if not os.path.isfile(full):
            continue
        if not PHOTO_PATTERN.match(f):
            unparseable.append(f)

    def sort_key(item):
        row = item["row_code"] or "ZZ"
        gc = item["grave_code"]
        try:
            n = int(gc)
        except (TypeError, ValueError):
            n = 10**9
        return (row, n, gc or "", item.get("plot_code", ""))

    needs_photo.sort(key=sort_key)
    link_fix.sort(key=sort_key)
    orphans.sort(key=lambda o: (o["row_letter"], int(o["grave"]) if o["grave"].lstrip('-').isdigit() else 10**9, o["plot"], o["filename"]))

    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Sheet 1: graves that need a photo taken in the field
    ws1 = wb.active
    ws1.title = "Needs Photo"
    ws1.append(["Row", "Grave Code", "Plot Code", "Name(s)", "Listed File Name (broken)"])
    style_header(ws1)
    for item in needs_photo:
        ws1.append([
            item["row_code"], item["grave_code"], item["plot_code"],
            item["names"], item["listed_file_name"],
        ])
    for col, w in zip("ABCDE", [8, 14, 14, 50, 28]):
        ws1.column_dimensions[col].width = w
    ws1.freeze_panes = "A2"

    # Sheet 2: photo exists on disk under correct name but register isn't linked
    ws2 = wb.create_sheet("Link Fix Only")
    ws2.append(["Row", "Grave Code", "Plot Code", "Name(s)", "Listed File Name", "Photo On Disk"])
    style_header(ws2)
    for item in link_fix:
        ws2.append([
            item["row_code"], item["grave_code"], item["plot_code"],
            item["names"], item["listed_file_name"], item["disk_match"],
        ])
    for col, w in zip("ABCDEF", [8, 14, 14, 50, 28, 28]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    # Sheet 3: photos on disk that don't match any grave in the register
    ws3 = wb.create_sheet("Orphan Photos")
    ws3.append(["Filename", "Parsed Row", "Parsed Grave", "Parsed Plot", "Note"])
    style_header(ws3)
    for o in orphans:
        ws3.append([o["filename"], o["row_letter"], o["grave"], o["plot"], "no matching grave in register"])
    for f in unparseable:
        ws3.append([f, "", "", "", "filename does not match the {row}{grave}{plot}.ext pattern"])
    for col, w in zip("ABCDE", [22, 12, 14, 12, 60]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    wb.save(OUT_PATH)

    total_graves = len(graves)
    x_rows = sum(1 for k in graves if k[0].upper() == "X")
    print(f"Total grave entries in register: {total_graves}  (excluding {x_rows} register-only X rows)")
    print(f"Needs a photo taken in the field: {len(needs_photo)}")
    print(f"Photo on disk, just needs xlsx link: {len(link_fix)}")
    print(f"Orphan photos on disk: {len(orphans)}  (+ {len(unparseable)} unparseable filenames)")
    print(f"Wrote: {OUT_PATH}")


if __name__ == "__main__":
    main()
